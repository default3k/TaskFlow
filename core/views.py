from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import IntegrityError
from django.utils import timezone
from django.http import HttpResponse
from django.db import models
from .models import User, Company, CompanyJoinRequest, Project, Task, TaskComment, Notification
from .forms import (
    RegistrationForm, CompanyCreateForm, CompanyJoinRequestForm,
    ProjectForm, TaskForm, TaskStatusForm, TaskCommentForm
)

# Импорты для Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ (ДОЛЖНЫ БЫТЬ В НАЧАЛЕ) ==========

def create_notification(user, notification_type, message, link=''):
    """Создать уведомление для пользователя"""
    Notification.objects.create(
        user=user,
        type=notification_type,
        message=message,
        link=link
    )


# ========== ДАЛЕЕ ВСЕ ТВОИ VIEW-ФУНКЦИИ ==========
# (register, user_login, user_logout, dashboard и т.д.)

def register(request):
    """Регистрация пользователя"""
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.username}!')
            return redirect('dashboard')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = RegistrationForm()
    
    return render(request, 'core/register.html', {'form': form})


def user_login(request):
    """Вход в систему"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.username}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Неверное имя пользователя или пароль')
    return render(request, 'core/login.html')


def user_logout(request):
    """Выход из системы"""
    logout(request)
    messages.info(request, 'Вы вышли из системы')
    return redirect('login')


@login_required
def dashboard(request):
    """Главная страница (дашборд)"""
    user = request.user
    context = {}
    
    if user.company:
        tasks = Task.objects.filter(company=user.company)
        context['tasks_total'] = tasks.count()
        context['tasks_done'] = tasks.filter(status='done').count()
        context['tasks_in_progress'] = tasks.filter(status='in_progress').count()
        context['tasks_review'] = tasks.filter(status='review').count()
        context['tasks_waiting'] = tasks.filter(status='waiting').count()
        context['my_tasks'] = tasks.filter(assigned_to=user).order_by('-created_at')[:10]
        context['projects'] = Project.objects.filter(company=user.company).order_by('-created_at')[:5]
        
        if user.can_manage_company():
            context['pending_requests'] = CompanyJoinRequest.objects.filter(
                company=user.company, status='pending'
            )
    else:
        context['my_requests'] = CompanyJoinRequest.objects.filter(user=user)
        context['available_companies'] = Company.objects.exclude(
    id__in=user.join_requests.filter(status__in=['pending', 'approved']).values('company_id')
    )
    
    context['unread_count'] = Notification.objects.filter(user=user, is_read=False).count()

    return render(request, 'core/dashboard.html', context)


@login_required
def company_create(request):
    """Создание компании"""
    if request.method == 'POST':
        form = CompanyCreateForm(request.POST)
        if form.is_valid():
            company = form.save(commit=False)
            company.save()
            user = request.user
            user.company = company
            user.role = 'owner'
            user.save()
            messages.success(request, f'Компания "{company.name}" создана!')
            return redirect('dashboard')
    else:
        form = CompanyCreateForm()
    return render(request, 'core/company_create.html', {'form': form})


@login_required
def company_join_request(request, company_id):
    """Подача заявки на вступление в компанию"""
    company = get_object_or_404(Company, id=company_id)
    user = request.user
    
    # Если пользователь уже в компании — выходим
    if user.company == company:
        messages.warning(request, 'Вы уже состоите в этой компании')
        return redirect('dashboard')
    
    # Проверяем, есть ли одобренная заявка
    approved = CompanyJoinRequest.objects.filter(
        user=user, company=company, status='approved'
    ).first()
    
    # Если пользователь был удален из компании, старая одобренная заявка должна быть удалена
    # Проверяем: если есть approved, но user.company != company, значит его удалили
    if approved and user.company != company:
        # Удаляем старую одобренную заявку, чтобы можно было подать новую
        approved.delete()
        approved = None
        messages.info(request, 'Вы можете подать новую заявку в компанию')
    
    # Если есть активная одобренная заявка (значит пользователь все еще в компании)
    if approved:
        messages.warning(request, 'Вы уже состоите в этой компании')
        return redirect('dashboard')
    
    # Если есть отклонённая заявка — удаляем её
    rejected = CompanyJoinRequest.objects.filter(
        user=user, company=company, status='rejected'
    ).first()
    if rejected:
        rejected.delete()
        messages.info(request, 'Старая отклонённая заявка удалена. Можете подать новую.')
    
    # Проверяем, нет ли уже pending заявки
    pending = CompanyJoinRequest.objects.filter(
        user=user, company=company, status='pending'
    ).first()
    if pending:
        messages.warning(request, 'Вы уже подали заявку в эту компанию (ожидает рассмотрения)')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CompanyJoinRequestForm(request.POST)
        if form.is_valid():
            join_request = form.save(commit=False)
            join_request.user = user
            join_request.company = company
            join_request.save()
            
            # Уведомляем владельца и админов
            owner = company.get_owner()
            if owner:
                create_notification(
                    owner,
                    'request_approved',
                    f'📝 Новая заявка! {user.username} хочет вступить в компанию "{company.name}"',
                    '/company/requests/'
                )
            admins = User.objects.filter(company=company, role='admin')
            for admin in admins:
                create_notification(
                    admin,
                    'request_approved',
                    f'📝 Новая заявка! {user.username} хочет вступить в компанию "{company.name}"',
                    '/company/requests/'
                )
            
            messages.success(request, f'Заявка в компанию "{company.name}" отправлена. Ожидайте подтверждения.')
            return redirect('dashboard')
    else:
        form = CompanyJoinRequestForm()
    
    return render(request, 'core/company_join_request.html', {'form': form, 'company': company})

@login_required
def company_requests_manage(request):
    """Управление заявками в компанию"""
    if not request.user.can_manage_company():
        messages.error(request, 'Нет прав')
        return redirect('dashboard')
    
    requests_list = CompanyJoinRequest.objects.filter(
        company=request.user.company, 
        status='pending'
    )
    
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        action = request.POST.get('action')
        join_request = get_object_or_404(CompanyJoinRequest, id=request_id, company=request.user.company)
        
        if action == 'approve':
            join_request.status = 'approved'
            join_request.save()
            user = join_request.user
            user.company = request.user.company
            user.role = 'developer'
            user.save()
            
            # Уведомление
            create_notification(
                user,
                'request_approved',
                f'Ваша заявка в компанию "{request.user.company.name}" одобрена! Ваша роль: разработчик',
            )
            messages.success(request, f'{user.username} принят в компанию как разработчик')
            
        elif action == 'reject':
            join_request.status = 'rejected'
            join_request.save()
            
            # Уведомление
            create_notification(
                join_request.user,
                'request_rejected',
                f'Ваша заявка в компанию "{request.user.company.name}" отклонена',
            )
            messages.info(request, f'Заявка {join_request.user.username} отклонена')
        
        return redirect('company_requests_manage')
    
    return render(request, 'core/company_requests_manage.html', {'requests': requests_list})


@login_required
def project_create(request):
    """Создание проекта"""
    if not request.user.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    if not request.user.can_manage_tasks():
        messages.error(request, 'У вас нет прав на создание проектов')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.company = request.user.company
            project.created_by = request.user
            project.save()
            messages.success(request, f'Проект "{project.name}" создан')
            return redirect('project_detail', project_id=project.id)
    else:
        form = ProjectForm()
    
    return render(request, 'core/project_form.html', {'form': form, 'title': 'Создание проекта'})


@login_required
def project_detail(request, project_id):
    """Детальная страница проекта"""
    project = get_object_or_404(Project, id=project_id, company=request.user.company)
    return render(request, 'core/project_detail.html', {'project': project})


@login_required
def task_create(request):
    """Создание задачи (владелец, админ, менеджер)"""
    if not request.user.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    if not request.user.can_manage_tasks():
        messages.error(request, 'У вас нет прав на создание задач')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = TaskForm(request.POST, company=request.user.company)
        if form.is_valid():
            task = form.save(commit=False)
            task.company = request.user.company
            task.created_by = request.user
            
            if task.assigned_to and task.assigned_to.company != request.user.company:
                messages.error(request, 'Исполнитель не состоит в вашей компании')
                return redirect('task_create')
            
            if task.project and task.project.company != request.user.company:
                messages.error(request, 'Проект принадлежит другой компании')
                return redirect('task_create')
            
            task.save()
            
            # Уведомление исполнителю
            if task.assigned_to and task.assigned_to != request.user:
                create_notification(
                    task.assigned_to,
                    'task_assigned',
                    f'Вам назначена задача "{task.title}" от {request.user.username}',
                    f'/task/{task.id}/'
                )
            
            messages.success(request, f'Задача "{task.title}" создана')
            if task.project:
                return redirect('project_detail', project_id=task.project.id)
            return redirect('tasks_list')
    else:
        form = TaskForm(company=request.user.company)
        project_id = request.GET.get('project')
        if project_id:
            try:
                project = Project.objects.get(id=project_id, company=request.user.company)
                form.fields['project'].initial = project
            except Project.DoesNotExist:
                pass
    
    return render(request, 'core/task_form.html', {'form': form, 'title': 'Создание задачи'})


@login_required
def tasks_list(request):
    """Список задач с поиском, сортировкой и фильтрацией"""
    if not request.user.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    tasks = Task.objects.filter(company=request.user.company)
    
    # ПОИСК
    search_query = request.GET.get('search', '')
    if search_query:
        tasks = tasks.filter(
            models.Q(title__icontains=search_query) |
            models.Q(description__icontains=search_query)
        )
    
    # ФИЛЬТРАЦИЯ
    status_filter = request.GET.get('status')
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    priority_filter = request.GET.get('priority')
    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)
    
    assigned_filter = request.GET.get('assigned')
    if assigned_filter == 'me':
        tasks = tasks.filter(assigned_to=request.user)
    elif assigned_filter == 'unassigned':
        tasks = tasks.filter(assigned_to__isnull=True)
    
    project_filter = request.GET.get('project')
    if project_filter:
        tasks = tasks.filter(project_id=project_filter)
    
    # СОРТИРОВКА
    sort_by = request.GET.get('sort', '-created_at')
    
    # Обработка сортировки по сроку (с учетом NULL)
    if sort_by == 'deadline':
        # Сначала NULL, потом даты (дальние сначала)
        tasks = tasks.order_by(models.F('deadline').asc(nulls_first=True))
    elif sort_by == '-deadline':
        # Сначала даты (ближайшие), потом NULL
        tasks = tasks.order_by(models.F('deadline').asc(nulls_last=True))
    else:
        # Обычная сортировка
        allowed_sorts = ['title', '-title', 'created_at', '-created_at', 
                         'priority', '-priority', 'status', '-status']
        if sort_by in allowed_sorts:
            tasks = tasks.order_by(sort_by)
        else:
            tasks = tasks.order_by('-created_at')
    
    projects = Project.objects.filter(company=request.user.company)
    
    context = {
        'tasks': tasks,
        'search_query': search_query,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'assigned_filter': assigned_filter,
        'project_filter': project_filter,
        'sort_by': sort_by,
        'projects': projects,
    }
    
    return render(request, 'core/tasks_list.html', context)

@login_required
def task_detail(request, task_id):
    """Детальная страница задачи с комментариями"""
    task = get_object_or_404(Task, id=task_id, company=request.user.company)
    comments = task.comments.all()
    
    if request.method == 'POST':
        # Проверяем, есть ли в POST данные комментария
        if 'text' in request.POST:
            form = TaskCommentForm(request.POST)
            if form.is_valid():
                comment = form.save(commit=False)
                comment.task = task
                comment.author = request.user
                comment.save()
                messages.success(request, 'Комментарий добавлен')
            else:
                messages.error(request, 'Ошибка при добавлении комментария')
        return redirect('task_detail', task_id=task.id)
    
    form = TaskCommentForm()
    return render(request, 'core/task_detail.html', {
        'task': task,
        'comments': comments,
        'form': form,
    })

@login_required
def task_update_status(request, task_id):
    """Обновление статуса задачи"""
    task = get_object_or_404(Task, id=task_id)
    user = request.user
    
    if task.company != user.company:
        messages.error(request, 'Нет доступа')
        return redirect('dashboard')
    
    can_change = False
    if user.can_change_any_status():
        can_change = True
    elif user.can_change_own_status() and task.assigned_to == user:
        can_change = True
    
    if not can_change:
        messages.error(request, 'Нет прав для изменения статуса')
        return redirect('task_detail', task_id=task.id)
    
    old_status = task.status
    new_status = request.POST.get('status')
    status_names = dict(Task.STATUS_CHOICES)
    
    if new_status in status_names:
        task.status = new_status
        if new_status == 'in_progress' and not task.started_at:
            task.started_at = timezone.now()
        if new_status == 'done':
            task.completed_at = timezone.now()
        task.save()
        
        # Уведомления
        if task.created_by != user:
            create_notification(
                task.created_by,
                'task_status_changed',
                f'Статус задачи "{task.title}" изменен: {status_names[old_status]} → {status_names[new_status]}',
                f'/task/{task.id}/'
            )
        if task.assigned_to and task.assigned_to != user and task.assigned_to != task.created_by:
            create_notification(
                task.assigned_to,
                'task_status_changed',
                f'Статус задачи "{task.title}" изменен: {status_names[old_status]} → {status_names[new_status]}',
                f'/task/{task.id}/'
            )
        
        messages.success(request, 'Статус обновлен')
    
    return redirect('task_detail', task_id=task.id)


@login_required
def task_delete(request, task_id):
    """Удаление задачи (владелец, админ, менеджер)"""
    task = get_object_or_404(Task, id=task_id)
    user = request.user
    
    if task.company != user.company:
        messages.error(request, 'Нет доступа к этой задаче')
        return redirect('dashboard')
    
    if not user.can_delete_tasks():
        messages.error(request, 'У вас нет прав на удаление задач')
        return redirect('task_detail', task_id=task.id)
    
    if task.project:
        project_id = task.project.id
        task_title = task.title
        task.delete()
        messages.success(request, f'Задача "{task_title}" удалена')
        return redirect('project_detail', project_id=project_id)
    
    task_title = task.title
    task.delete()
    messages.success(request, f'Задача "{task_title}" удалена')
    return redirect('tasks_list')


@login_required
def company_members(request):
    """Управление сотрудниками компании"""
    if not request.user.can_manage_roles():
        messages.error(request, 'У вас нет прав для управления ролями сотрудников')
        return redirect('dashboard')
    
    members = User.objects.filter(company=request.user.company).exclude(role='applicant')
    pending_requests = CompanyJoinRequest.objects.filter(company=request.user.company, status='pending')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        member_id = request.POST.get('member_id')
        
        if action == 'change_role':
            new_role = request.POST.get('new_role')
            member = get_object_or_404(User, id=member_id, company=request.user.company)
            
            if request.user.role == 'owner':
                member.role = new_role
                member.save()
                messages.success(request, f'Роль {member.username} изменена на {member.get_role_display()}')
            elif request.user.role == 'admin' and member.role != 'owner':
                member.role = new_role
                member.save()
                messages.success(request, f'Роль {member.username} изменена на {member.get_role_display()}')
            else:
                messages.error(request, f'Нельзя изменить роль пользователя {member.username}')
        
        elif action == 'remove_member':
            member = get_object_or_404(User, id=member_id, company=request.user.company)
            
            if member.role == 'owner':
                messages.error(request, 'Нельзя удалить владельца компании')
            else:
                company = request.user.company
                company_name = company.name
                
                # Удаляем одобренную заявку пользователя
                CompanyJoinRequest.objects.filter(
                    user=member, 
                    company=company, 
                    status='approved'
                ).delete()
                
                member.company = None
                member.role = 'applicant'
                member.save()

                create_notification(
                    member,
                    'request_rejected',
                    f'Вас удалили из компании "{company_name}"'
                )
                
                messages.success(request, f'{member.username} удален из компании')
        
        return redirect('company_members')
    
    return render(request, 'core/company_members.html', {
        'members': members,
        'pending_requests': pending_requests,
    })


@login_required
def company_edit(request):
    """Редактирование компании"""
    if not request.user.can_manage_company():
        messages.error(request, 'У вас нет прав')
        return redirect('dashboard')
    
    if request.method == 'POST':
        company = request.user.company
        company.name = request.POST.get('name')
        company.description = request.POST.get('description')
        company.save()
        messages.success(request, 'Информация о компании обновлена')
        return redirect('company_members')
    
    return render(request, 'core/company_edit.html', {'company': request.user.company})


@login_required
def company_leave(request):
    """Выход из компании (любой сотрудник, кроме владельца)"""
    user = request.user
    
    if not user.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    if user.role == 'owner':
        messages.error(request, 'Владелец не может выйти из компании. Используйте "Удалить компанию"')
        return redirect('company_members')
    
    if request.method == 'POST':
        company_name = user.company.name
        user.company = None
        user.role = 'applicant'
        user.save()
        
        messages.success(request, f'Вы вышли из компании "{company_name}"')
        return redirect('dashboard')
    
    return render(request, 'core/company_confirm_leave.html', {'company': user.company})


@login_required
def company_delete(request):
    """Удаление компании (только владелец)"""
    user = request.user
    
    if not user.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    if user.role != 'owner':
        messages.error(request, 'Только владелец может удалить компанию')
        return redirect('dashboard')
    
    company = user.company
    
    if request.method == 'POST':
        company_name = company.name
        company.delete()
        
        user.company = None
        user.role = 'applicant'
        user.save()
        
        messages.success(request, f'Компания "{company_name}" полностью удалена')
        return redirect('dashboard')
    
    return render(request, 'core/company_confirm_delete.html', {'company': company})


@login_required
def export_tasks_excel(request):
    """Экспорт задач в Excel"""
    if not request.user.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    tasks = Task.objects.filter(company=request.user.company)
    
    status_filter = request.GET.get('status')
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    assigned_filter = request.GET.get('assigned')
    if assigned_filter == 'me':
        tasks = tasks.filter(assigned_to=request.user)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Задачи"
    
    headers = [
        'ID', 'Название', 'Статус', 'Приоритет', 'Исполнитель', 
        'Создатель', 'Проект', 'Дата создания', 'Срок выполнения'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.id)
        ws.cell(row=row, column=2, value=task.title)
        ws.cell(row=row, column=3, value=task.get_status_display())
        ws.cell(row=row, column=4, value=task.get_priority_display())
        ws.cell(row=row, column=5, value=task.assigned_to.username if task.assigned_to else 'Не назначен')
        ws.cell(row=row, column=6, value=task.created_by.username)
        ws.cell(row=row, column=7, value=task.project.name if task.project else 'Без проекта')
        ws.cell(row=row, column=8, value=task.created_at.strftime('%d.%m.%Y %H:%M') if task.created_at else '')
        ws.cell(row=row, column=9, value=task.deadline.strftime('%d.%m.%Y %H:%M') if task.deadline else '')
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    filename = f'tasks_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@login_required
def profile_view(request):
    """Личный профиль пользователя"""
    user = request.user
    
    if user.company:
        my_tasks = Task.objects.filter(company=user.company, assigned_to=user)
        tasks_total = my_tasks.count()
        tasks_done = my_tasks.filter(status='done').count()
        tasks_in_progress = my_tasks.filter(status='in_progress').count()
        
        total_time = 0
        for task in my_tasks.filter(status='done'):
            time_spent = task.get_time_spent()
            if time_spent:
                total_time += time_spent
    else:
        tasks_total = tasks_done = tasks_in_progress = total_time = 0
    
    context = {
        'user': user,
        'tasks_total': tasks_total,
        'tasks_done': tasks_done,
        'tasks_in_progress': tasks_in_progress,
        'total_time': round(total_time, 1),
    }
    return render(request, 'core/profile.html', context)


@login_required
def profile_edit(request):
    """Редактирование профиля"""
    user = request.user
    
    if request.method == 'POST':
        user.email = request.POST.get('email')
        user.phone = request.POST.get('phone')
        user.position = request.POST.get('position')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        
        if User.objects.exclude(id=user.id).filter(email=user.email).exists():
            messages.error(request, 'Этот email уже используется')
            return redirect('profile_edit')
        
        user.save()
        
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if new_password:
            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, 'Пароль успешно изменен. Войдите снова.')
                return redirect('login')
            else:
                messages.error(request, 'Пароли не совпадают')
                return redirect('profile_edit')
        
        messages.success(request, 'Профиль успешно обновлен')
        return redirect('profile')
    
    return render(request, 'core/profile_edit.html', {'user': user})


@login_required
def notifications_view(request):
    """Страница уведомлений"""
    notifications = request.user.notifications.all()
    
    if request.method == 'POST':
        notification_id = request.POST.get('mark_read')
        if notification_id:
            notif = get_object_or_404(Notification, id=notification_id, user=request.user)
            notif.is_read = True
            notif.save()
        elif 'mark_all_read' in request.POST:
            request.user.notifications.update(is_read=True)
        return redirect('notifications')
    
    return render(request, 'core/notifications.html', {'notifications': notifications})

@login_required
def profile_view_public(request, user_id):
    """Просмотр профиля другого пользователя (для админов и владельцев)"""
    viewer = request.user
    
    # Проверяем, что смотрящий состоит в компании
    if not viewer.company:
        messages.error(request, 'Вы не состоите в компании')
        return redirect('dashboard')
    
    # Получаем целевого пользователя
    target_user = get_object_or_404(User, id=user_id)
    
    # Проверяем, что целевой пользователь в той же компании
    if target_user.company != viewer.company:
        messages.error(request, 'Этот пользователь не состоит в вашей компании')
        return redirect('dashboard')
    
    # Права на просмотр: только админ или владелец могут смотреть чужие профили
    if not viewer.can_manage_company():
        # Обычный пользователь может смотреть только свой профиль
        if viewer.id != target_user.id:
            messages.error(request, 'У вас нет прав на просмотр профилей других сотрудников')
            return redirect('dashboard')
    
    # Собираем статистику по задачам пользователя
    my_tasks = Task.objects.filter(company=viewer.company, assigned_to=target_user)
    tasks_total = my_tasks.count()
    tasks_done = my_tasks.filter(status='done').count()
    tasks_in_progress = my_tasks.filter(status='in_progress').count()
    tasks_review = my_tasks.filter(status='review').count()
    tasks_waiting = my_tasks.filter(status='waiting').count()
    
    # Вычисляем общее время
    total_time = 0
    for task in my_tasks.filter(status='done'):
        time_spent = task.get_time_spent()
        if time_spent:
            total_time += time_spent
    
    # Получаем последние задачи пользователя
    recent_tasks = my_tasks.order_by('-created_at')[:5]
    
    context = {
        'profile_user': target_user,
        'tasks_total': tasks_total,
        'tasks_done': tasks_done,
        'tasks_in_progress': tasks_in_progress,
        'tasks_review': tasks_review,
        'tasks_waiting': tasks_waiting,
        'total_time': round(total_time, 1),
        'recent_tasks': recent_tasks,
        'is_own_profile': (viewer.id == target_user.id),
        'can_manage': viewer.can_manage_company(),
    }
    
    return render(request, 'core/profile_public.html', context)